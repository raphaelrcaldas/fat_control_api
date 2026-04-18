"""Geração de planilha Excel profissional para etapas."""

from datetime import date, datetime, time
from io import BytesIO

# Paleta de cores (stdlib, sem openpyxl — seguras no nível do módulo).
_BLUE_DARK = '1F3864'
_BLUE_MED = '2E5E9E'
_BLUE_LIGHT = 'D6E4F0'
_BLUE_ZEBRA = 'F2F6FC'
_BORDER_CLR = 'D0D5DD'
_WHITE = 'FFFFFF'

# Placeholders para openpyxl. Inicializados como None para que ruff
# (F821) enxergue os nomes no nível do módulo; o import real e a
# construção das constantes acontece em _init_openpyxl(), chamado
# no começo de generate_etapas_xlsx(). Assim openpyxl (~200ms) só
# carrega quando o endpoint de export é efetivamente invocado — fora
# do cold start.
Workbook = None
get_column_letter = None
Alignment = None
Border = None
Side = None
Font = None
PatternFill = None
_TITLE_FONT = None
_TITLE_FILL = None
_META_FONT = None
_META_FILL = None
_HEADER_FONT = None
_HEADER_FILL = None
_HEADER_BORDER = None
_DATA_BORDER = None
_ZEBRA_FILL = None
_TOTAL_FILL = None
_TOTAL_FONT = None
_CENTER = None
_CENTER_WRAP = None
_LEFT_WRAP = None
_RIGHT = None

_lazy_initialized = False


def _init_openpyxl() -> None:
    """Importa openpyxl e popula as constantes de estilo do módulo.

    Idempotente — chamadas subsequentes são no-op. Deve ser chamada
    no começo de toda função pública que use os estilos.
    """
    global _lazy_initialized  # noqa: PLW0603
    if _lazy_initialized:
        return

    from openpyxl import Workbook  # noqa: PLC0415
    from openpyxl.styles import (  # noqa: PLC0415
        Alignment,
        Border,
        Font,
        PatternFill,
        Side,
    )
    from openpyxl.utils import get_column_letter  # noqa: PLC0415

    # Escopo local: os nomes importados acima são locais a esta função;
    # globals().update() sobrescreve os placeholders `None` no módulo.
    globals().update({
        'Workbook': Workbook,
        'get_column_letter': get_column_letter,
        'Alignment': Alignment,
        'Border': Border,
        'Side': Side,
        'Font': Font,
        'PatternFill': PatternFill,
        '_TITLE_FONT': Font(bold=True, size=14, color=_WHITE),
        '_TITLE_FILL': PatternFill(
            start_color=_BLUE_DARK,
            end_color=_BLUE_DARK,
            fill_type='solid',
        ),
        '_META_FONT': Font(size=10, italic=True, color='4472C4'),
        '_META_FILL': PatternFill(
            start_color=_BLUE_LIGHT,
            end_color=_BLUE_LIGHT,
            fill_type='solid',
        ),
        '_HEADER_FONT': Font(bold=True, size=10, color=_WHITE),
        '_HEADER_FILL': PatternFill(
            start_color=_BLUE_MED,
            end_color=_BLUE_MED,
            fill_type='solid',
        ),
        '_HEADER_BORDER': Border(
            top=Side(style='thin', color=_BLUE_DARK),
            bottom=Side(style='thin', color=_BLUE_DARK),
            left=Side(style='thin', color=_BLUE_DARK),
            right=Side(style='thin', color=_BLUE_DARK),
        ),
        '_DATA_BORDER': Border(
            top=Side(style='thin', color=_BORDER_CLR),
            bottom=Side(style='thin', color=_BORDER_CLR),
            left=Side(style='thin', color=_BORDER_CLR),
            right=Side(style='thin', color=_BORDER_CLR),
        ),
        '_ZEBRA_FILL': PatternFill(
            start_color=_BLUE_ZEBRA,
            end_color=_BLUE_ZEBRA,
            fill_type='solid',
        ),
        '_TOTAL_FILL': PatternFill(
            start_color=_BLUE_LIGHT,
            end_color=_BLUE_LIGHT,
            fill_type='solid',
        ),
        '_TOTAL_FONT': Font(bold=True, size=10),
        '_CENTER': Alignment(horizontal='center', vertical='center'),
        '_CENTER_WRAP': Alignment(
            horizontal='center',
            vertical='center',
            wrap_text=True,
        ),
        '_LEFT_WRAP': Alignment(
            horizontal='left',
            vertical='center',
            wrap_text=True,
        ),
        '_RIGHT': Alignment(horizontal='right', vertical='center'),
    })
    _lazy_initialized = True


def _min_to_hhmm(minutes: int) -> str:
    """Converte minutos para formato HH:MM."""
    return f'{minutes // 60:02d}:{minutes % 60:02d}'


def _fmt_time(t: time | None) -> str:
    """Formata time para HH:MM."""
    if t is None:
        return '-'
    return t.strftime('%H:%M')


def _fmt_date(d: date | None) -> str:
    """Formata date para DD/MM/YYYY."""
    if d is None:
        return '-'
    return d.strftime('%d/%m/%Y')


def _fmt_val(val, fmt: str = 'int') -> str | int | float:
    """Formata valor ou retorna '-' se nulo."""
    if val is None:
        return '-'
    if fmt == 'float':
        return float(val)
    return int(val)


def _auto_width(
    ws,
    col_count: int,
    header_row: int,
    wide_cols: frozenset[str] | None = None,
    skip_cols: frozenset[str] | None = None,
):
    """Ajusta largura das colunas baseado no conteudo."""
    if wide_cols is None:
        wide_cols = frozenset()
    if skip_cols is None:
        skip_cols = frozenset()
    for col_idx in range(1, col_count + 1):
        max_len = 0
        col_letter = get_column_letter(col_idx)
        hdr_val = str(ws.cell(row=header_row, column=col_idx).value or '')
        if hdr_val in skip_cols:
            continue
        is_wide = hdr_val in wide_cols
        for row in ws.iter_rows(
            min_row=header_row,
            max_row=ws.max_row,
            min_col=col_idx,
            max_col=col_idx,
        ):
            for cell in row:
                if cell.value is None:
                    continue
                val = str(cell.value)
                lines = val.split('\n')
                longest = max(len(ln) for ln in lines)
                if cell.row == header_row:
                    longest = int(longest * 1.3) + 3
                max_len = max(max_len, longest)
        # Colunas com texto longo: fator extra
        if is_wide:
            max_len = int(max_len * 1.2)
        width = max(max_len + 2, 8)
        ws.column_dimensions[col_letter].width = width


# Colunas com wrap_text a esquerda
_WRAP_HEADERS = frozenset({
    'Esforço Aéreo',
    'Tripulantes',
})
# Colunas com wrap_text centralizado
_WRAP_CENTER_HEADERS = frozenset({
    'Cod OI',
    'Regime',
})
# Colunas numericas alinhadas a direita
_NUM_HEADERS = frozenset({
    'Pousos',
    'TOW',
    'Comb',
    'Lub',
})
# Colunas centralizadas
_CENTER_HEADERS = frozenset({
    'PAX',
    'Carga',
})


def generate_etapas_xlsx(
    etapas: list,
    oi_data: dict | None,
    trip_data: dict | None,
    columns: dict[str, bool],
) -> BytesIO:
    """Gera planilha Excel profissional de etapas.

    Args:
        etapas: lista de rows (Etapa model instances)
        oi_data: OIEtapaOut agrupados por etapa_id
        trip_data: TripEtapaOut agrupados por etapa_id
        columns: flags de colunas opcionais
    """
    _init_openpyxl()
    wb = Workbook()
    ws = wb.active
    ws.title = 'Etapas'

    # ── Montar colunas dinamicamente ───────────────
    headers = [
        'Data',
        'Origem',
        'Destino',
        'DEP',
        'ARR',
        'TV',
        'Aeronave',
    ]

    # esforco_aereo agora gera 3 colunas separadas
    optional = [
        ('pousos', ['Pousos']),
        ('nivel', ['Nível']),
        ('tow', ['TOW']),
        ('pax', ['PAX']),
        ('carga', ['Carga']),
        ('comb', ['Comb']),
        ('lub', ['Lub']),
        (
            'esforco_aereo',
            [
                'Cod OI',
                'Esforço Aéreo',
                'D/N/V',
            ],
        ),
        ('tripulantes', ['Tripulantes']),
    ]
    active_optional: list[tuple[str, list[str]]] = []
    for flag, hdrs in optional:
        if columns.get(flag, False):
            active_optional.append((flag, hdrs))
            headers.extend(hdrs)

    col_count = len(headers)
    last_col = get_column_letter(col_count)

    # ── Linha 1: Titulo institucional ──────────────
    ws.merge_cells(f'A1:{last_col}1')
    title_cell = ws['A1']
    title_cell.value = '1º/1º GT — Relatório de Etapas'
    title_cell.font = _TITLE_FONT
    title_cell.fill = _TITLE_FILL
    title_cell.alignment = Alignment(
        horizontal='center',
        vertical='center',
    )
    ws.row_dimensions[1].height = 35

    for col_idx in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = _TITLE_FILL

    # ── Linha 2: Metadados ─────────────────────────
    total_tvoo = sum(e.tvoo for e in etapas)
    now = datetime.now()
    meta_text = (
        f'Exportado em: {now.strftime("%d/%m/%Y %H:%M")}'
        f'  |  Total de etapas: {len(etapas)}'
        f'  |  TV total: {_min_to_hhmm(total_tvoo)}'
    )
    ws.merge_cells(f'A2:{last_col}2')
    meta_cell = ws['A2']
    meta_cell.value = meta_text
    meta_cell.font = _META_FONT
    meta_cell.fill = _META_FILL
    meta_cell.alignment = Alignment(
        horizontal='center',
        vertical='center',
    )
    ws.row_dimensions[2].height = 22

    for col_idx in range(1, col_count + 1):
        cell = ws.cell(row=2, column=col_idx)
        cell.fill = _META_FILL

    # ── Linha 3: Header da tabela ──────────────────
    header_row = 3
    for col_idx, hdr in enumerate(headers, start=1):
        cell = ws.cell(
            row=header_row,
            column=col_idx,
            value=hdr,
        )
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.border = _HEADER_BORDER
        cell.alignment = _CENTER_WRAP
    ws.row_dimensions[header_row].height = 28

    ws.auto_filter.ref = f'A{header_row}:{last_col}{header_row}'

    # ── Dados ──────────────────────────────────────
    data_start = header_row + 1
    sum_tvoo = 0
    sum_pousos = 0
    sum_pax = 0
    sum_carga = 0
    sum_comb = 0
    sum_lub = 0.0

    for row_idx, etapa in enumerate(etapas):
        r = data_start + row_idx
        is_even = row_idx % 2 == 0

        values = [
            _fmt_date(etapa.data),
            etapa.origem.upper(),
            etapa.destino.upper(),
            _fmt_time(etapa.dep),
            _fmt_time(etapa.arr),
            _min_to_hhmm(etapa.tvoo),
            etapa.anv,
        ]
        sum_tvoo += etapa.tvoo

        for flag, _ in active_optional:
            if flag == 'pousos':
                v = _fmt_val(etapa.pousos)
                if etapa.pousos is not None:
                    sum_pousos += etapa.pousos
                values.append(v)
            elif flag == 'nivel':
                v = etapa.nivel if etapa.nivel else '-'
                values.append(v)
            elif flag == 'tow':
                values.append(_fmt_val(etapa.tow))
            elif flag == 'pax':
                v = _fmt_val(etapa.pax)
                if etapa.pax is not None:
                    sum_pax += etapa.pax
                values.append(v)
            elif flag == 'carga':
                v = _fmt_val(etapa.carga)
                if etapa.carga is not None:
                    sum_carga += etapa.carga
                values.append(v)
            elif flag == 'comb':
                v = _fmt_val(etapa.comb)
                if etapa.comb is not None:
                    sum_comb += etapa.comb
                values.append(v)
            elif flag == 'lub':
                v = _fmt_val(etapa.lub, 'float')
                if etapa.lub is not None:
                    sum_lub += float(etapa.lub)
                values.append(v)
            elif flag == 'esforco_aereo':
                ois = oi_data.get(etapa.id, []) if oi_data else []
                if ois:
                    cod_lines = [oi.tipo_missao_cod for oi in ois]
                    esf_lines = [oi.esf_aer for oi in ois]
                    reg_map = {
                        'd': 'D',
                        'n': 'N',
                        'v': 'V',
                    }
                    reg_lines = [
                        reg_map.get(oi.reg, oi.reg.upper()) for oi in ois
                    ]
                    values.append('\n'.join(cod_lines))
                    values.append('\n'.join(esf_lines))
                    values.append('\n'.join(reg_lines))
                else:
                    values.extend(['-', '-', '-'])
            elif flag == 'tripulantes':
                trips = trip_data.get(etapa.id, []) if trip_data else []
                if trips:
                    trigs = [t.trig.upper() for t in trips]
                    values.append(' / '.join(trigs))
                else:
                    values.append('-')

        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(
                row=r,
                column=col_idx,
                value=val,
            )
            cell.border = _DATA_BORDER

            if is_even:
                cell.fill = _ZEBRA_FILL

            hdr_name = headers[col_idx - 1]
            if hdr_name in _WRAP_HEADERS:
                cell.alignment = _LEFT_WRAP
            elif hdr_name in _WRAP_CENTER_HEADERS:
                cell.alignment = _CENTER_WRAP
            elif hdr_name in _NUM_HEADERS:
                cell.alignment = _RIGHT
            elif hdr_name in _CENTER_HEADERS:
                cell.alignment = _CENTER
            else:
                cell.alignment = _CENTER

        # Auto-height: ajusta pela celula com mais linhas
        max_lines = 1
        for val in values:
            if isinstance(val, str) and '\n' in val:
                max_lines = max(
                    max_lines,
                    val.count('\n') + 1,
                )
        if max_lines > 1:
            ws.row_dimensions[r].height = max_lines * 15

    # ── Linha de totais ────────────────────────────
    total_row = data_start + len(etapas)
    for col_idx in range(1, col_count + 1):
        cell = ws.cell(row=total_row, column=col_idx)
        cell.fill = _TOTAL_FILL
        cell.font = _TOTAL_FONT
        cell.border = Border(
            top=Side(style='medium', color=_BLUE_MED),
            bottom=Side(
                style='medium',
                color=_BLUE_MED,
            ),
            left=Side(
                style='thin',
                color=_BORDER_CLR,
            ),
            right=Side(
                style='thin',
                color=_BORDER_CLR,
            ),
        )
        cell.alignment = _CENTER

    ws.cell(row=total_row, column=1).value = 'TOTAL'
    ws.cell(
        row=total_row,
        column=1,
    ).alignment = _CENTER

    ws.cell(
        row=total_row,
        column=6,
    ).value = _min_to_hhmm(sum_tvoo)

    # Totais das opcionais
    base_count = 7
    col_offset = 0
    for flag, hdrs in active_optional:
        col = base_count + col_offset + 1
        cell = ws.cell(row=total_row, column=col)
        if flag == 'pousos':
            cell.value = sum_pousos
            cell.alignment = _RIGHT
        elif flag == 'pax':
            cell.value = sum_pax
            cell.alignment = _RIGHT
        elif flag == 'carga':
            cell.value = sum_carga
            cell.alignment = _RIGHT
        elif flag == 'comb':
            cell.value = sum_comb
            cell.alignment = _RIGHT
        elif flag == 'lub':
            cell.value = round(sum_lub, 1)
            cell.alignment = _RIGHT
        col_offset += len(hdrs)

    # ── Layout e impressao ─────────────────────────
    ws.freeze_panes = 'A4'

    # Larguras fixas para colunas base (conteudo previsivel)
    _FIXED_WIDTHS = {
        'Data': 12,
        'Origem': 9,
        'Destino': 9,
        'DEP': 7,
        'ARR': 7,
        'TV': 7,
        'Aeronave': 10,
        'Pousos': 8,
        'Nível': 7,
        'TOW': 7,
        'PAX': 6,
        'Carga': 8,
        'Comb': 7,
        'Lub': 6,
        'Cod OI': 9,
        'Regime': 9,
    }
    for col_idx, hdr in enumerate(headers, start=1):
        fixed = _FIXED_WIDTHS.get(hdr)
        if fixed:
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = fixed

    # Auto-width apenas para colunas dinamicas
    _auto_width(
        ws,
        col_count,
        header_row=header_row,
        wide_cols=_WRAP_HEADERS,
        skip_cols=frozenset(_FIXED_WIDTHS.keys()),
    )
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = '1:3'
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    # ── Salvar ─────────────────────────────────────
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
